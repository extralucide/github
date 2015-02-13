#!/usr/bin/env python 2.7.3
# -*- coding: utf-8 -*-
__author__ = 'olivier'
from openpyxl import load_workbook,Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import re
from check_llr import CheckLLR
from datetime import datetime
try:
    from Tkinter import *
    ##    import Tkinter              # Python 2
    import ttk
except ImportError:
    from tkinter import *
from tool import Tool
from synergy import Synergy
import time
from math import floor
from os.path import join
from openpyxl.styles import Font
from openpyxl.styles.borders import BORDER_THIN

class Derived(CheckLLR,Synergy):

    def __init__(self,
                 basename="",
                 hlr_selected=False,
                 **kwargs):

        for key in kwargs:
            self.__dict__[key] = kwargs[key]

        if "session_started" in self.__dict__:
            Synergy.__init__(self,self.session_started)

        if "general_output_txt" in self.__dict__:
            CheckLLR.__init__(self,
                              basename,
                              hlr_selected,
                              general_output_txt=self.__dict__["general_output_txt"])
        else:
            CheckLLR.__init__(self,
                              basename,
                              hlr_selected)

        self.index_row = 0
        self.index_column = 0
        self.log_filename = None
        self.log_handler = None

    def countDerived(self,type="SWRD"):
        self.tbl_req_derived=[]
        if type in self.dico_specifications:
            derived_value =  self.dico_specifications[type]["derived"]
        else:
            derived_value = "YES"
        #if type == "SSCS":
        nb = self.FilterReq({"derived":derived_value},self.tbl_req_derived)
        #else:
        #    nb = self.FilterReq({"derived":"YES"},self.tbl_req_derived)
        self.log( "Nb requirements derived found:{:d}".format(nb),gui_display=True)

    def export(self):
        wb = load_workbook(filename = 'template/clean_saq345_derived_requirement_review.xlsx')
        #ws = wb['Register']
        ws = wb.get_sheet_by_name(name = 'Register')
        self.putLogo(ws)
        #wb = Workbook(True)
        #ws = wb.create_sheet()
        # Title
        ws.cell('C7').value = ""
        # Reference
        ws.cell('C8').value = ""
        # Issue
        ws.cell('C9').value = ""
        row = 14
        for req in self.tbl_req_derived:
            line = []
            value = self.tbl_list_llr[req]
            # Create list from dictionary
            file = self.list_llr_vs_file[str(req)][0]
            line.append(file)
            line.append(req)
            body = CheckLLR.getAtribute(value,"body")
            #row.append(Tool.removeNonAscii(body))
            line.append(body)
            rationale = CheckLLR.getAtribute(value,"rationale")
            #row.append(Tool.removeNonAscii(rationale))
            line.append(rationale)
            line.append("B")
            self.sqlite_connect()
            chapter = self.sqlite_get(req)
            line.append(chapter)
            row += 1
            for col_idx in range(1,7):
                column = get_column_letter(col_idx)
                ws.cell('%s%s'%(column, row)).value = '%s' % (line[col_idx - 1])
            #ws.append(row)
        CheckLLR.set_border(ws, "A12:P12")
        CheckLLR.set_border(ws, "F2:J4")
        CheckLLR.set_border(ws,
                        "C15:C%s"%(row),
                        font = Font(name='Arial',size=10,bold=False),
                        border_style=BORDER_THIN,
                        alignment_horizontal="left")
        CheckLLR.set_border(ws,
                        "D15:D%s"%(row),
                        font = Font(name='Arial',size=10,bold=False),
                        border_style=BORDER_THIN,
                        alignment_horizontal="left")
        # save the file
        filename = "Derived_Req_Feedback_%d.xlsx" % floor(time.time())
        wb.save(join("result",filename))
        return filename

if __name__ == '__main__':
    dirname = "/Users/olivier/github/local/WHCC"
    dirname = "C:\\Documents and Settings\\appereo1\\Bureau\\sqa\\ENM\\SWRD"
    hlr = Derived(dirname,hlr_selected=True)
    list_dir = hlr.listDir()
    hlr.invert()
    hlr.getDerived()
    hlr.export()


