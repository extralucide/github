#!/usr/bin/env python 2.7.3
# -*- coding: utf-8 -*-
import Queue
from Tkconstants import END, NORMAL, CURRENT
from math import floor
import re
import threading
import time
from export_doc import BuildDoc
from reviews import Review
from synergy import Synergy
from tool import Tool
from ccb import CCB
import os
from os.path import join
import webbrowser
from openpyxl import load_workbook,Workbook
from openpyxl.compat import range
from openpyxl.styles import Style,Font,PatternFill,Border,Side,Alignment
from openpyxl.styles.borders import BORDER_THIN,BORDER_MEDIUM
from check_llr import CheckLLR
from check_is import CheckIS
from export_derived import Derived
from get_ig_jquery import easyIG,getQA
from datetime import datetime
from conf import VERSION
from convert_xml_html import HtmlConverter
import sys
sys.path.append("intelhex")
from intelhex import IntelHex,IntelHex16bit

__author__ = 'olivier'
# TODO: Attention a _getParentCR appellee par getParentCR dans la classe Synergy
# TODO: Il faut simplifier
class ThreadQuery(threading.Thread,Synergy):
    def lock(self):
        # global count_baseline
        # count_baseline +=1
##        print "Wait lock release: " + str(count_baseline) + "\n"
##        print "amount of threads alive:" + str(threading.active_count()) + "\n"
        self.verrou.acquire()
    def unlock(self):
        self.verrou.release()
##        print "Release lock.\n"

    def __init__(self,
                 name_id="",
                 master=None,
                 queue="",
                 login="",
                 password="",
                 **kwargs):
        """
        :param name_id:
        :param master:
        :param queue:
        :param kwargs:
        :return:
        """

        for key in kwargs:
            self.__dict__[key] = kwargs[key]
        if "no_start_session" not in self.__dict__:
            self.no_start_session = False

        threading.Thread.__init__(self)
        # Create the queue
        self.queue = queue
        self.master_ihm = master
        Synergy.__init__(self,
                         session_started=False,
                         ihm=self.master_ihm)
        self.login = login
        self.password = password
        self.running = 1
        self.database = None
        if "system" not in self.__dict__:
            print "Missing system definition."
            self.system = ""
        if "item" in self.__dict__:
            # Get database name and aircraft name
            self.database,self.aircraft = self.get_sys_item_database(self.system,
                                                                     self.item)
        else:
            print "Missing item definition."
            self.item = ""
        self.component = self.master_ihm.component
        if self.database is None:
            self.database,self.aircraft = self.get_sys_database()
        self.author = ""
        self.reference = ""
        self.release = ""
        self.project = ""
        self.baseline = ""
        self.revision = ""
        # Recursive lock
        self.verrou = threading.RLock()
        self.name_id = name_id
        self.input_data_filter = ""
        self.peer_reviews_filter = ""
        self.list_projects = []

        # Display system name
        self.master_ihm.log("System: {:s}".format(self.system),False)
        # Display item name
        self.master_ihm.log("Item: {:s}".format(self.item),False)
        # Display configuration item ID
        ci_id = self.get_ci_sys_item_identification(self.system,self.item)
        if ci_id is not None:
            self.master_ihm.log("CI ID: {:s}".format(ci_id),False)
        else:
            self.master_ihm.log("CI ID: Unknown",False)

        self.export_cr_list_filename = "export_CR_list_template.xlsx"
        self.dico_cr_log = {}
        self.dico_cr_transition = {}
        self.easyig = easyIG()
        self.getqa = getQA()

    def stopSession(self):
        if self.session_started:
            stdout,stderr = self.ccm_query('stop','Stop Synergy session')
            if stdout != "":
                # remove \r
                text = re.sub(r"\r\n",r"\n",stdout)
                self.master_ihm.log(text,False)
            if stderr:
                 # remove \r
                text = re.sub(r"\r\n",r"\n",stderr)
                self.master_ihm.log(text,False)

    def processIncoming(self):
        """
        Handle all the messages currently in the queue (if any).
         - BUILD_DOCX
            . Store selection
         - START_SESSION
         - GET_BASELINES
         - GET_RELEASES
         - GET_PROJECTS
         - etc.
        """
        while self.queue.qsize():
            try:
                self.lock()
##                print threading.enumerate();
                # Check contents of message
                action = self.queue.get(0)
                print time.strftime("%H:%M:%S", time.localtime()) + " Commmand: " + action
                if action == "BUILD_CID":
                    data = self.queue.get(1)
                    print "TEST_DATA",data
                    self.release = data[0]
                    self.project = data[1]
                    self.baseline = data[2]
                    release = data[0]
                    baseline = data[2]
                    project = data[1]
                    implemented = data[3]
                    item = data[4]
                    previous_baseline = data[5]
                    detect = data[6]
                    cr_type = data[7]
                    component = data[8]
                    cr_domain = data[9]
                    self.storeSelection(self.project,
                                        self.system,
                                        self.release,
                                        self.baseline)
                    self.build_doc_thread = threading.Thread(None,self._generateCID,None,(release,
                                                                                          baseline,
                                                                                          project,
                                                                                          implemented,
                                                                                          item,
                                                                                          previous_baseline,
                                                                                          detect,
                                                                                          cr_type,
                                                                                          component,
                                                                                          cr_domain))
                    self.build_doc_thread.start()
                elif action == "EASY_IG":
                    self.master_ihm.log("Launch easyIG.")
                    self.easy_ig_thread = threading.Thread(None,self._easyIG,None)
                    self.easy_ig_thread.start()
                elif action == "GET_QA_ACTIONS":
                    url_root = self.queue.get(1)
                    name,mail,tel,service,qams_user_id = self.get_user_infos(self.login)
                    self.get_qa_thread = threading.Thread(None,self._getQA,None,(qams_user_id,url_root,name))
                    self.get_qa_thread.start()
                elif action == "READ_BPROC":
                    bproc_filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    xsl = join(current_dir,"template\\xsl_procedure_ece-1.0.xsl")
                    html_name = Tool.getFileName(bproc_filename)
                    html_filename = join(current_dir,"result\\" + html_name)
                    self.read_bproc_thread = threading.Thread(None,self._readBPROC,None,(bproc_filename,xsl,html_filename))
                    self.read_bproc_thread.start()
                elif action == "READ_GPROC":
                    gproc_filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    xsl = join(current_dir,"template\\gproc_makefile.xsl")
                    html_name = Tool.getFileName(gproc_filename)
                    html_filename = join(current_dir,"result\\" + html_name)
                    self.read_gproc_thread = threading.Thread(None,self._readBPROC,None,(gproc_filename,xsl,html_filename))
                    self.read_gproc_thread.start()
                elif action == "READ_RTP":
                    filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    xsl = join(current_dir,"template\\rtp.xsl")
                    html_name = Tool.getFileName(filename)
                    html_filename = join(current_dir,"result\\" + html_name)
                    self.read_gproc_thread = threading.Thread(None,self._readBPROC,None,(filename,xsl,html_filename))
                    self.read_gproc_thread.start()
                elif action == "READ_EOC":
                    eoc_filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    if self.config_parser.has_section("EOC"):
                        addr_hw_sw_compatibility = self.getOptions("EOC","addr_hw_sw_compatibility")
                        addr_pn = self.getOptions("EOC","addr_pn")
                        addr_checksum = self.getOptions("EOC","addr_checksum")
                        addr_hw_sw_compatibility_range = addr_hw_sw_compatibility.split(",")
                        addr_pn_range = addr_pn.split(",")
                        addr_checksum_range = addr_checksum.split(",")
                        dico_addr={"hw_sw_compat":addr_hw_sw_compatibility_range,
                                   "pn":addr_pn_range,
                                   "checksum":addr_checksum_range}
                    else:
                        dico_addr={"hw_sw_compat":("0x400","0x402"),
                                   "pn":("0x400","0x424"),
                                   "checksum":("0x4DE8","0x4DEA")}
                    self.read_eoc_thread = threading.Thread(None,self.thread_readEOC,None,(eoc_filename,dico_addr))
                    self.read_eoc_thread.start()
                elif action == "GET_BASELINE_STATUS":
                    baseline = self.queue.get(1)
                    if baseline != "":
                        self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("GET_BASELINE_STATUS","",baseline))
                        self.send_cmd_thread.start()
                elif action == "GET_RELEASE_INFO":
                    release = self.queue.get(1)
                    if release != "":
                        self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("GET_RELEASE_INFO",release))
                        self.send_cmd_thread.start()
                elif action == "BUILD_SQAP":
                    data = self.queue.get(1)
                    author = data[0]
                    self.reference = data[1]
                    self.revision = data[2]
                    self.build_doc_thread = threading.Thread(None,self._generateSQAP,None,(author,self.reference,self.revision,self.aircraft,self.system,self.item))
                    self.build_doc_thread.start()
                elif action == "BUILD_CCB":
                    data = self.queue.get(1)
                    dico_parameters = data[0]
                    dico_parameters["login"] = self.login
                    cr_with_parent = data[1]
                    cr_workflow = data[2]
                    cr_domain = data[3]
                    log_on = data[4]
                    list_cr_for_ccb = data[5]
                    status_list = data[6]
                    ccb_time = data[7]
                    self.build_doc_thread = threading.Thread(None,self._generateCCB,None,(dico_parameters,
                                                                                          cr_with_parent,
                                                                                          cr_workflow,
                                                                                          cr_domain,
                                                                                          log_on,
                                                                                          list_cr_for_ccb,
                                                                                          status_list,
                                                                                          ccb_time))
                    self.build_doc_thread.start()

                elif action == "BUILD_REVIEW_REPORT":
                    review_id = self.queue.get(1)
                    empty = self.queue.get(2)
                    self.build_doc_thread = threading.Thread(None,self._generateReviewReport,None,(review_id,empty))
                    self.build_doc_thread.start()
                elif action == "BUILD_DELIVERY_SHEET":
                    type_sds = self.queue.get(1)
                    dico_tags = self.queue.get(2)
                    #print dico_tags
                    self.build_doc_thread = threading.Thread(None,self._generateDeliverySheet,None,(type_sds,dico_tags))
                    self.build_doc_thread.start()
                elif action == "START_SESSION":
                    # start synergy session
                    self.start_session_thread = threading.Thread(None,self._startSession,None,(self.system,
                                                                                               self.item,
                                                                                               self.database,
                                                                                               self.login,
                                                                                               self.password,
                                                                                               self.aircraft))
                    self.start_session_failed = False
                    self.start_session_thread.start()
                    self.launch_session = True
                    self.setSessionStarted()
                elif action == "GET_BASELINES":
                    if self.session_started:
                        #release = self.master_ihm.release
                        release = self.queue.get(1)
                        self.get_baselines_thread = threading.Thread(None,self._getBaselinesList,None,(release,))
                        self.get_baselines_thread.start()
                elif action == "GET_RELEASES":
                    if self.session_started:

                        regexp = self.queue.get(1)
                        active = self.master_ihm.getActive()
                        if active:
                            query = "release -active -u -l"
                        else:
                            query = "release -u -l"
                        self.master_ihm.log("ccm " + query)
                        self.get_releases_thread = threading.Thread(None,self._getReleasesList,None,(query,regexp))
                        self.get_releases_thread.start()
                elif action == "GET_PROJECTS":
                    if self.session_started:
                        baseline = self.master_ihm.baseline
                        release = self.master_ihm.release
                        query = self._defineProjectQuery(release,
                                                         baseline)
                        self.master_ihm.log("ccm " + query)
                        self.get_projects_thread = threading.Thread(None,self._getProjectsList,None,(query,release,baseline))
                        self.get_projects_thread.start()
                elif action == "READ_STATUS":
                    self.set_status_thread = threading.Thread(None,self._getSessionStatus,None)
                    self.set_status_thread.start()
                elif action == "CLOSE_SESSION":
                    self.set_status_thread = threading.Thread(None,self._closeSession,None)
                    self.set_status_thread.start()
                elif action == "MAKE_DIFF":
                    data = self.queue.get(1)
                    baseline_prev = data[0]
                    baseline_cur = data[1]
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("BASELINES_DIFF","","","",baseline_prev,baseline_cur))
                    self.send_cmd_thread.start()
                elif action == "SHOW_BASELINE":
                    data = self.queue.get(1)
                    baseline_cur = data[0]
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("BASELINES_SHOW","","","","",baseline_cur))
                    self.send_cmd_thread.start()
                elif action == "SEND_CMD":
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None)
                    self.send_cmd_thread.start()
                elif action == "EXPORT_CR":
                    cr_id = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._exportCR,None,(cr_id,))
                    self.send_cmd_thread.start()
                elif action == "LIST_ITEMS":
                    release = self.queue.get(1)
                    project = self.queue.get(2)
                    baseline = self.queue.get(3)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("ITEMS",release,baseline,project))
                    self.send_cmd_thread.start()
                elif action == "SCOPE":
                    release = self.queue.get(1)
                    project = self.queue.get(2)
                    baseline = self.queue.get(3)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("SCOPE",release,baseline,project))
                    self.send_cmd_thread.start()
                elif action == "LIST_TASKS":
                    release = self.queue.get(1)
                    baseline = self.queue.get(2)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("TASKS",release,baseline))
                    self.send_cmd_thread.start()
                elif action == "LIST_HISTORY":
                    release = self.queue.get(1)
                    baseline = self.queue.get(2)
                    project = self.queue.get(3)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("HISTORY",release,baseline,project))
                    self.send_cmd_thread.start()
                elif action == "GET_RELEASE_VS_BASELINE":
                    if self.session_started:
                        baseline = self.queue.get(1)
                        self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("GET_RELEASE_VS_BASELINE","",baseline))
                        self.send_cmd_thread.start()
                elif action == "EXEC_USER_CMD":
                        tbl_user_cmd = self.queue.get(1)
                        self.send_cmd_thread = threading.Thread(None,self._execUserCmd,None,(tbl_user_cmd,))
                        self.send_cmd_thread.start()
                elif action == "PREVIEW_CR_QUERY":
                    self.send_cmd_thread = threading.Thread(None,self._preview_CR_Query,None)
                    self.send_cmd_thread.start()

                elif action == "GET_CR":
                    data = self.queue.get(1)
                    baseline = data[0]
                    ccb_type = data[1]
                    extension = True
                    for_review_on = data[2]
                    cr_with_parent = data[3]
                    log_on = data[4]
                    component_type = data[5]
                    detected_on = data[6]
                    implemented_for = data[7]
                    old_cr_workflow = data[8]
                    ccb_time = data[9]
                    self.build_doc_thread = threading.Thread(None,self._getCR,None,(baseline,
                                                                                    extension,
                                                                                    for_review_on,
                                                                                    cr_with_parent,
                                                                                    log_on,
                                                                                    component_type,
                                                                                    detected_on,
                                                                                    implemented_for,
                                                                                    old_cr_workflow,
                                                                                    ccb_time))
                    self.build_doc_thread.start()

                elif action == "START_APACHE":
                    config= "httpd_ece.conf"
                    self.send_cmd_thread = threading.Thread(None,self.__apache_start,None,(config,))
                    self.send_cmd_thread.start()

                elif action == "CHECK_UPDATE":
                    self.check_update_thread = threading.Thread(None,self._checkUpdate,None,)
                    self.check_update_thread.start()

                elif action == "RELOAD_CONFIG":
                    # Get config
##                    self.__loadConfig()
##                    interface.log("Config file docid.ini reloaded.")
                    pass
                elif action == "RELOAD_BASELINEBOX":
                    if self.session_started:
                        stdout = self.queue.get(1)
                        if stdout != "":
                            self.master_ihm.baselinelistbox.configure(bg="white")
                            self.master_ihm.log("Available baseline found:")
                            output = stdout.splitlines()
                            self.master_ihm.baselinelistbox.delete(0, END)
                            if len(output) > 1:
                                self.master_ihm.baselinelistbox.insert(END, "All")
                                self.master_ihm.baselinelistbox.selection_set(first=0)
                            for line in output:
                                line = re.sub(r"^ *[0-9]{1,3}\) ",r"",line)
                                self.master_ihm.baselinelistbox.insert(END, line)
                                self.master_ihm.baselinelistbox_1.insert(END, line)
                                self.master_ihm.baselinelistbox_2.insert(END, line)
                                self.master_ihm.log(line)
                            self.master_ihm.releaselistbox.selection_set(first=0)
                            self.master_ihm.baselinelistbox.configure(bg="white")
                        else:
                            self.master_ihm.resetBaselineListbox()
                            self.master_ihm.log(" No available baselines found.")
                        #self.resetProjectListbox()
                        self.master_ihm.baselinelistbox.configure(state=NORMAL)
                        # Set scrollbar at the bottom
                        self.master_ihm.defill()
                elif action == "RELOAD_RELEASEBOX":
                    if self.session_started:
                        stdout = self.queue.get(1)
                        if stdout != "":
                            output = stdout.splitlines()
                            # Populate release listbox
                            self.master_ihm.updateReleaseListBox(output)
                        else:
                            self.master_ihm.noneReleaseListBox()
                elif action == "RELOAD_PROJECTBOX":
                    if self.session_started:
                        stdout = self.queue.get(1)
                        release = self.queue.get(2)
                        baseline_selected = self.queue.get(3)
                        if stdout != "":
                            #self.master_ihm.projectlistbox.delete(0, END)
                            self.master_ihm.projectlistbox.clear()
                            output = stdout.splitlines()
                            # Here the list of projects is set
                            self.list_projects = []
                            if Tool.isAttributeValid(baseline_selected):
                                if Tool.isAttributeValid(release):
                                    for line in output:
                                        line = re.sub(r"^ *[0-9]{1,3}\) ",r"",line)
                                        m = re.match(r'(.*)-(.*);(.*|<void>)$',line)
                                        if m:
                                            project = m.group(1) + "-" + m.group(2)
                                            baseline_string = m.group(3)
                                            baseline_splitted = baseline_string.split(',')
                                            for baseline in baseline_splitted:
                                                baseline = re.sub(r".*#",r"",baseline)
                                                if baseline == baseline_selected:
                                                    self.list_projects.append(project)
                                                    break
                                        else:
                                            m = re.match(r'^Baseline(.*):$',line)
                                            if not m:
                                                project = line
                                                self.list_projects.append(project)
                                else:
                                    num = 0
                                    for project in output:
                                        if num > 0:
                                            self.list_projects.append(project)
                                        num += 1
                            else:
                                for line in output:
                                    line = re.sub(r"^ *[0-9]{1,3}\) ",r"",line)
                                    m = re.match(r'(.*)-(.*);(.*)$',line)
                                    if m:
                                        project = m.group(1) + "-" + m.group(2)
                                        #print "name " + m.group(1) + " version " + m.group(2)
                                    else:
                                        project = line
                                    self.list_projects.append(project)
                            # Update list of project of GUI
                            self.master_ihm.updateProjectListBox(self.list_projects)
                        else:
                            self.master_ihm.noneProjectListBox()
                elif action == "RELOAD_CRLISTBOX":
                    if self.session_started:
                        try:
                            print "Display CR RELOAD_CRLISTBOX"
                            list_cr = self.queue.get(1)
                            # Update list of project of GUI
                            crlistbox = self.master_ihm.crlistbox
                            crlistbox.configure(state=NORMAL)
                            crlistbox.delete(0, END)
                            inter = 0
                            for cr_description in list_cr:
                                crlistbox.insert(END, cr_description)
                                if inter % 2 == 0:
                                    crlistbox.itemconfig(inter,{'bg':'gray88','fg':'black'})
                                else:
                                    crlistbox.itemconfig(inter,{'bg':'lightgrey','fg':'black'})
                                inter += 1
                            crlistbox.configure(bg="white")
                        except AttributeError:
                            pass
                elif action == "CHECK_LLR":
                    dirname = self.queue.get(1)
                    hsid_dirname = self.queue.get(2)
                    self.send_cmd_thread = threading.Thread(None,self._checkLLRCmd,None,(dirname,False,("SWDD",),hsid_dirname))
                    self.send_cmd_thread.start()
                elif action == "CHECK_HLR":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._checkLLRCmd,None,(dirname,True,("SWRD","PLDRD")))
                    self.send_cmd_thread.start()
                elif action == "CHECK_UPPER":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._checkUpperCmd,None,(dirname,))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_IS_HLR":
                    dirname_req = self.queue.get(1)
                    dirname_upper = self.queue.get(2)
                    reference = self.queue.get(3)
                    issue = self.queue.get(4)
                    release = self.queue.get(5)
                    self.send_cmd_thread = threading.Thread(None,self._exportIS,None,(dirname_req,dirname_upper,True,reference,issue,release))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_IS_LLR":
                    dirname_req = self.queue.get(1)
                    dirname_upper = self.queue.get(2)
                    reference = self.queue.get(3)
                    issue = self.queue.get(4)
                    release = self.queue.get(5)
                    hsid_dirname = self.queue.get(6)
                    self.send_cmd_thread = threading.Thread(None,self._exportIS,None,(dirname_req,dirname_upper,False,reference,issue,release,hsid_dirname))
                    self.send_cmd_thread.start()
                elif action == "CHECK_IS_HLR":
                    dirname_upper = self.queue.get(1)
                    dirname_req = self.queue.get(2)
                    filename_is = self.queue.get(3)
                    component = self.queue.get(4)
                    self.send_cmd_thread = threading.Thread(None,self._checkISCmd,None,(dirname_upper,dirname_req,filename_is,component,True))
                    self.send_cmd_thread.start()
                elif action == "CHECK_IS_LLR":
                    dirname_upper = self.queue.get(1)
                    dirname_req = self.queue.get(2)
                    filename_is = self.queue.get(3)
                    component = ""
                    self.send_cmd_thread = threading.Thread(None,self._checkISCmd,None,(dirname_upper,dirname_req,filename_is,component))
                    self.send_cmd_thread.start()
                elif action == "CHECK_IS_DOC":
                    filename_is = self.queue.get(1)
                    cr_process_version = self.queue.get(2)
                    self.send_cmd_thread = threading.Thread(None,self._checkISDocCmd,None,(filename_is,cr_process_version))
                    self.send_cmd_thread.start()
                elif action == "GEN_DERIVED_HLR":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._genHLRDerivedCmd,None,(dirname,))
                    self.send_cmd_thread.start()
                elif action == "GEN_DERIVED_LLR":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._genLLRDerivedCmd,None,(dirname,))
                    self.send_cmd_thread.start()
                elif action == "GEN_DERIVED_UPPER":
                    dirname = self.queue.get(1)
                    upper = CheckLLR()
                    list_upper = upper.getListUpper()
                    self.send_cmd_thread = threading.Thread(None,self._genHLRDerivedCmd,None,(dirname,list_upper))
                    self.send_cmd_thread.start()
                elif action == "UPDATE_CHAPTER_HLR":
                    filename = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._getChapterDialogHLR,None,(filename,))
                    self.send_cmd_thread.start()
                else:
                    pass
                self.unlock()
            except Queue.Empty:
                pass

    def thread_readEOC(self,eoc_filename,dico_addr):
        hw_sw_compatibility,part_number,checksum = self._readEOC(eoc_filename,dico_addr)
        self.displayEOC_Info((hw_sw_compatibility,
                              part_number,
                              checksum))

    def periodicCall(self):
        """
        Check every 1000 ms if there is something new in the queue.
        """
##        print time.strftime("%H:%M:%S", time.localtime())
##        print time.strftime("PERIODIC CALL " + self.name_id)
        self.processIncoming()
        if not self.running:
            # This is the brutal stop of the system. You may want to do
            # some cleanup before actually shutting it down.
            import sys
            sys.exit(1)
        try:
            self.master_ihm.after(1000, self.periodicCall)
        except AttributeError:
            time.sleep(1)
            self.periodicCall

    def _setRelease(self):
        self.master_ihm.release = self.previous_release
        self.master_ihm.button_list_items.configure(state=NORMAL)
        self.master_ihm.button_list_tasks.configure(state=NORMAL)
        self.master_ihm.button_set_baselines.configure(state=NORMAL)
        self.master_ihm.setBaseline(self.master_ihm.release)

    def _setBaseline(self):
        self.master_ihm.baseline = self.previous_baseline
        self.master_ihm.setBaselineSynergy(self.master_ihm.baseline)
        self.master_ihm.projectlistbox.configure(state=NORMAL)
        self.master_ihm.button_find_projects.configure(state=NORMAL)
        self.master_ihm.button_list_items.configure(state=NORMAL)
        self.master_ihm.button_list_tasks.configure(state=NORMAL)
        executed = self._sendCmd("GET_RELEASE_VS_BASELINE","",self.master_ihm.baseline)
        if executed:
            pass
##            interface.button_select.configure(state=NORMAL)
        query = self._defineProjectQuery(self.master_ihm.release,self.master_ihm.baseline)
        self._getProjectsList(query,self.master_ihm.release,self.master_ihm.baseline)

    def _setProject(self):
        self.master_ihm.project = self.previous_project
        self.master_ihm.button_select.configure(state=NORMAL)
        self.master_ihm.button_create_delivery_sheet.configure(state=NORMAL)
        self.master_ihm.button_list_items.configure(state=NORMAL)
        self.master_ihm.button_list_tasks.configure(state=NORMAL)
        self.master_ihm.setProject(self.master_ihm.project)

    def _add(self, action):
        # add an action to the manager.  returns tags to use in
        # associated text widget
        tag = "hlink-%d" % len(self.links)
        self.links[tag] = action
        return "hlink", tag

    def _click(self, event):
        for tag in self.master_ihm.general_output_txt.tag_names(CURRENT):
            if tag[:6] == "hlink-":
                            self.links[tag]()

    def _startSession(self,
                      system,
                      item,
                      database,
                      login,
                      password,
                      aircraft="",
                      queue_thread_gui=None):
        """ Function to start Synergy session
             - invoke command ccm start ...
             - display synergy feedback
             - retrieve last session information
             - enable SELECT and REFRESH buttons
             - get list of releases
            called by the thread """
        # GUI/CLI
        print "_startSession"
        #self.master_ihm.put_in_gui_queue("Test on_main_thread")
        #self.lock()

        #self.previous_release = ""
        #self.previous_baseline = ""
        #self.previous_project = ""
        session_started = False
        if database is not None \
                and login != "":
            try:
                query = "start /nogui /q /d /usr/local/ccmdb/{:s} /u /usr/local/ccmdb/{:s} /s {:s} /n {:s} /pw {:s}".format(database,database,self.ccm_server,login,password)
                stdout,stderr = self.ccm_query(query,"Synergy session start")
            except UnicodeEncodeError:
                stdout = False
            self.master_ihm.resultStartSession(stdout,stderr)

        else:
            self.master_ihm.sayNoDatabase()
            self.start_session_failed = True
            stdout = ""

        return stdout

    def _easyIG(self):
        filename = self.easyig.get()
        self.master_ihm.displayHyperlink("hlink",filename,"Web page created.")
        self.easyig.start()

    def _getQA(self,qams_user_id,url_root,name=""):
        filename = self.getqa.get(qams_user_id,url_root=url_root,name=name)
        self.master_ihm.displayHyperlink("hlink_local_qams",filename,"Local web page created.")
        self.getqa.start()

    def _getReleasesList(self,query="cmd release -u -l",regexp=""):
        ''' get releases list '''
        self.lock()
        stdout,stderr = self.ccm_query(query,"Get releases")
        if regexp != "":
            output = stdout.splitlines()
            list_release = ""
            if stdout != "":
                for line in output:
                    m = re.match(regexp,line)
                    if m:
                        list_release += line + "\n"
            if list_release == "":
                self.master_ihm.log("Check release_regexp parameter in docid.ini which value is: " + regexp)
        else:
            list_release = stdout
        ##self.master_ihm.resultGenerateCID(docx_filename,
        #                                  False,
        #                                  text="SYNERGY GET RELEASES COMMAND")
        self.queue.put("RELOAD_RELEASEBOX") # action to reload release listbox
        self.queue.put(list_release)
        self.unlock()

    def _getBaselinesList(self,release):
        """
        get baseline list for Release/Baseline/Project window
        :param release:
        :return:
        """
        self.lock()
        if Tool.isAttributeValid(release):
            query = 'baseline -l -release {:s} -f "%name"'.format(release)
        else:
            query = 'baseline -l -f "%name"'
        self.master_ihm.log("ccm " + query)
        stdout,stderr = self.ccm_query(query,"Get baselines")
        if stdout != "":
            output = stdout.splitlines()
            self.master_ihm.updateBaselineListBox(output)
        else:
            self.master_ihm.noneBaselineListBox()

        # Set scrollbar at the bottom
        #self.master_ihm.defill()
        #self.master_ihm.success.config(fg='magenta',bg = 'green',text="COMMAND SUCCEEDED")
        self.unlock()

    def _getProjectsList(self,query,release,baseline_selected):
        """
        :param query:
        :param release:
        :param baseline_selected:
        :return:
        """
        self.lock()
        stdout,stderr = self.ccm_query(query,"Get projects")
        self.queue.put("RELOAD_PROJECTBOX") # action to get projects
        self.queue.put(stdout)
        self.queue.put(release)
        self.queue.put(baseline_selected)
        #self.master_ihm.success.config(fg='magenta',bg = 'green',text="COMMAND SUCCEEDED")
        self.unlock()

    def _exportCR(self,cr_id):
        """
        Function to export CR to web browser
        called by crlistbox_onselect
        """
        query = "query -t problem \"(problem_number='" + cr_id + "')\" -u -f \
                 \"<table border='1'>\
                 <cell name='CR_domain'>%CR_domain</cell>\
                 <cell name='CR_type'>%CR_type</cell>\
                 <cell name='crstatus'>%crstatus</cell>\
                 <cell name='problem_synopsis'>%problem_synopsis</cell>\
                 <cell name='SCR_In_Analysis_id'>%SCR_In_Analysis_id</cell>\
                 <cell name='create_time'>%create_time</cell>\
                 <cell name='CR_ECE_classification'>%CR_ECE_classification</cell>\
                 <cell name='CR_customer_classification'>%CR_customer_classification</cell>\
                 <cell name='CR_request_type'>%CR_request_type</cell>\
                 <cell name='CR_detected_on'>%CR_detected_on</cell>\
                 <cell name='CR_applicable_since'>%CR_applicable_since</cell>\
                 <cell name='CR_implemented_for'>%CR_implemented_for</cell>\
                 <cell name='CR_origin'>%CR_origin</cell>\
                 <cell name='CR_origin_desc'>%CR_origin_desc</cell>\
                 <cell name='CR_expected'>%CR_expected</cell>\
                 <cell name='CR_observed'>%CR_observed</cell>\
                 <cell name='CR_functional_impact'>%CR_functional_impact</cell>\
                 <cell name='CR_analysis'>%CR_analysis</cell>\
                 <cell name='CR_correction_description'>%CR_correction_description</cell>\
                 <cell name='CR_product_impact'>%CR_product_impact</cell>\
                 <cell name='CR_doc_impact'>%CR_doc_impact</cell>\
                 <cell name='CR_verif_impact'>%CR_verif_impact</cell>\
                 <cell name='impact_analysis'>%impact_analysis</cell>\
                 <cell name='functional_limitation_desc'>%functional_limitation_desc</cell>\
                 <cell name='implemented_modification'>%implemented_modification</cell>\
                 <cell name='CR_implementation_baseline'>%CR_implementation_baseline</cell>\
                 <cell name='SCR_Verif_Test_Bench'>%SCR_Verif_Test_Bench</cell>\
                 <cell name='SCR_Verif_Test_Procedure'>%SCR_Verif_Test_Procedure</cell>\
                 <cell name='CR_verification_activities'>%CR_verification_activities</cell>\
                 <cell name='functional_limitation'>%functional_limitation</cell>\
                 <cell name='SCR_Closed_id'>%SCR_Closed_id</cell>\
                 <cell name='SCR_Closed_time'>%SCR_Closed_time</cell>\
                 <cell name='problem_number'>%problem_number</cell>\
                 <cell name='modify_time'>%modify_time</cell>\
                 <cell name='SCR_Fixed_time'l>%SCR_Fixed_time</cell>\
                 </table>\""
##                 <cell name='transition_log'>%transition_log</cell>\
        executed = True
        filename = "log_SCR_" + cr_id + "_%d.html" % floor(time.time())
        #with open(self.gen_dir + filename, 'w') as of:
        if query != "":
            ccm_query = 'ccm ' + query + '\n'
            cmd_out = self._ccmCmd(query,False)
            # Replace STX and ETS and e cute characters
            char = {r'\x02':r'<',r'\x03':r'>',r'\xe9':r'e'}
            for before, after in char.iteritems():
                cmd_out = re.sub(before,after,cmd_out)
            if cmd_out == "":
                self.master_ihm.log("No result.")
                executed = False
            #
            # Get transition log
            #
            query = "query -t problem \"(problem_number='{:s}')\" -u -f \"%transition_log\"".format(cr_id)
            ccm_query = 'ccm ' + query + '\n'
            transi_log = self._ccmCmd(query,False)
            transi_log_filtered = self._filterASCII(transi_log)

            # Get parent CR
            tbl_parent_cr_id = self._getParentCR(cr_id)
            if tbl_parent_cr_id:
                #
                # Get parent ID information
                #
                parent_cr = ""
                for parent_cr_id in tbl_parent_cr_id:
                    res_parent_cr = self._getParentInfo(parent_cr_id)
                    if res_parent_cr:
                        parent_cr += res_parent_cr
                        self.master_ihm.log("Parent CR:" + res_parent_cr)
                    #else:
                    #    self.master_ihm.log("No result for _getParentInfo (twice).")
            else:
                parent_cr = "<td><IMG SRC=\"../img/changeRequestIcon.gif\">---</td><td>---</td><td>---</td><td>---</td><td>---</td>"
            self._parseCR(cmd_out,
                          transi_log_filtered,
                          parent_cr,
                          join(self.gen_dir,filename))
            # Get information CR
            #TODO:
        if executed:
            self.master_ihm.log("Command executed.")
            self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
            url = join(os.getcwd(),"result")
            url = join(url,filename)
            print "URL",url
            webbrowser.open(url)
            self.master_ihm.success.config(fg='magenta',bg = 'green',text="EXPORT SUCCEEDED")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="EXPORT FAILED")
        # Set scrollbar at the bottom
        self.master_ihm.defill()
        return executed

    def _getItems(self,release="",baseline="",project=""):
        global session_started
        output = ""
        output_format = "csv"
        release_name = re.sub(r"\/",r"",release)
        executed = False
        filename = "log_items_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
        filename += ".{:s}".format(output_format)
        if output_format == "csv":
            display_attr = ' -f "%release;%name;%version;%modify_time;%status;%task;%task_status;%change_request;%type" '
            show_header = "-nch"
        else:
            display_attr = ' -f "%release %name %version %modify_time %status %task %task_status %change_request %type" '
            show_header = "-ch"
        if baseline not in ("","All"):
            # Baseline
            # sh: show
            #  u: no number
            query = "baseline -sh objects  {:s} -u {:s}".format(baseline,display_attr)
            executed = True
        elif release not in ("","All"):
            # Query with a specifcic release
            #  ch: Specifies to use column headers for the output
            # nch: Specifies not to use column headers for the output
            #   u: no number
            query = 'query -sby name {:s} -n *.* -u -release {:s} '.format(show_header,release)
            if project not in ("*","All",""):
                # a project is selected
                # get sub-projects
                name, version = self.getProjectInfo(project)
                query += '"recursive_is_member_of(cvtype=\'project\' and name=\'{:s}\' and version=\'{:s}\' , \'none\')" {:s}'.format(name,version)
            query += display_attr
            executed = True
        elif project not in ("","All"):
            # No baseline, nor release selected but a project is
            query = 'query -sby name {:s} -n *.* -u "(is_member_of(\'{:s}\'))" {:s}'.format(show_header,release,display_attr)
            executed = True
        else:
            self.master_ihm.log("Please select a release or a baseline or a project.")
        if executed:
            self.master_ihm.log(" ccm " + query)
            self.master_ihm.defill()
            ccm_query = 'ccm ' + query + '\n\n'
            self.master_ihm.log("List objects (directories and executable objects are discarded).")
            cmd_out = self._ccmCmd(query)
            with open(join(self.gen_dir,filename), 'w') as of:
                if output_format == "csv":
                    header = "Release;Name;Version;Modify time;Status;Task;Task status;CR;Type\n"
                    of.write(header)
                else:
                    of.write(ccm_query)
                output = cmd_out.splitlines()
                for line in output:
                    # Skip directory or relocatable objects
                    # Skip automatic tasks and components tasks
                    # Remove Baseline info at the beginning
                    if output_format == "csv":
                        if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                            # For CLI
                            print line
                            of.write(line)
                            of.write("\n")
                    else:
                        if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line):
                            of.write(line)
                            of.write("\n")
        if executed:
            self.master_ihm.log("Command executed.")
            try:
                self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
            except AttributeError:
                pass
        # Set scrollbar at the bottom
        self.master_ihm.defill()
        return output

    def _execUserCmd(self,tbl_user_cmd):
        executed = True
        export = True
        filename = "log_%d.txt" % floor(time.time())
        with open(join(self.gen_dir,filename), 'w') as of:
            for query in tbl_user_cmd:
                if query != "":
                    self.master_ihm.log('ccm ' + query)
                    ccm_query = 'ccm ' + query + '\n'
                    cmd_out = self._ccmCmd(query)
                    if cmd_out == None:
                        executed = False
                        break
                    try:
                        of.write(ccm_query)
                    except UnicodeEncodeError as exception:
                        print "Character not supported:", exception
                    of.write(cmd_out)
                    if cmd_out == "":
                        self.master_ihm.log("No result.")
                        executed = False
        if executed:
            self.master_ihm.log("Command executed.")
            if export:
                 # Create hyperlink
                if filename is not None:
                    #self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
                    self.master_ihm.resultGenerateCID(filename,
                                                      False,
                                                      text="USER COMMAND")
                else:
                    self.master_ihm.resultGenerateCID(False,
                                                      False,
                                                      text="USER COMMAND")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="COMMAND FAILED")

    def _sendCmd(self,cmd="",
                 release="",
                 baseline="",
                 project="",
                 baseline_prev="",
                 baseline_cur=""):
        try:
            self.master_ihm.success.config(fg='red',bg = 'yellow',text="SYNERGY COMMAND IN PROGRESS")
        except AttributeError:
            pass
        export = True
        global session_started
        output_format = "csv"
        release_name = re.sub(r"\/",r"",release)
        filename = "log_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
        if output_format == "txt":
            filename += ".txt"
        else:
            filename += ".csv"
        executed = False
        if cmd == "SCOPE":
            if release not in ("","All") and project not in ("*","All",""):
                test_string = "SQAP"
                text_found = ""
                project_name, project_version = self.getProjectInfo(project)
                query = "finduse -query \"release='" + release + "' and (cvtype='xls' or cvtype='doc' or cvtype='pdf' or cvtype='ascii' or cvtype='csrc') and recursive_is_member_of(cvtype='project' and name='"+ project_name +"' and version='"+ project_version +"' , 'none')\""
                self.master_ihm.log('ccm ' + query)
                self.master_ihm.general_output_txt.see(END)
                ccm_query = 'ccm ' + query + '\n\n'
                cmd_out = self._ccmCmd(query)
                output = cmd_out.splitlines()
##                test_string_1 = "Input_Data"
##                test_string_2 = "SQAP"
                list_items_skipped_1 = []
                list_items_skipped_2 = []
                regexp_1 = '^(.*)'+ project_name + '\\\\' + re.escape(self.input_data_filter) + '\\\\(.*)-(.*)@(.*)-(.*)$'
                regexp_2 = '^(.*)'+ project_name + '\\\\' + re.escape(self.peer_reviews_filter) + '\\\\(.*)-(.*)@(.*)-(.*)$'
                for line in output:
##                    print "Tested: " + line
                    # ex: SW_PLAN\SDP\IS_SDP_SW_PLAN_SQA.xlsm-1.7.0@SW_PLAN-1.3
                    m = re.match(regexp_1,line)
                    if m:
##                        text_found = m.group(2)
                        list_items_skipped_1.append(m.group(2))
                    else:
                        pass
                    m = re.match(regexp_2,line)
                    if m:
##                        text_found = m.group(2)
                        list_items_skipped_2.append(m.group(2))
                    else:
                        pass
##                print regexp_1
##                print regexp_2
                list_wo_doublons_1 = list(set(list_items_skipped_1))
                list_wo_doublons_2 = list(set(list_items_skipped_2))
##                print list_wo_doublons_1
##                print list_wo_doublons_2
                executed = True
            else:
                self.master_ihm.log("Please select a release and a project.")
        elif cmd == "GET_BASELINE_STATUS":
            executed = self._getBaselineInfo(baseline)
            export = False
        elif cmd == "GET_RELEASE_INFO":
            executed = self._getReleaseInfo(release)
            export = False
        elif cmd == "ITEMS":
            filename = "log_items_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
            if output_format == "csv":
                display_attr = ' -f "%release;%name;%version;%modify_time;%status;%task;%task_status;%change_request;%type" '
                show_header = "-nch"
                filename += ".csv"
            else:
                display_attr = ' -f "%release %name %version %modify_time %status %task %task_status %change_request %type" '
                show_header = "-ch"
                filename += ".txt"
            if baseline not in ("","All"):
                # Baseline
                # sh: show
                #  u: no number
                query = 'baseline -sh objects  ' + baseline + " -u "
                query += display_attr
                executed = True
            elif release not in ("","All"):
                # Query with a specifcic release
                #  ch: Specifies to use column headers for the output
                # nch: Specifies not to use column headers for the output
                #   u: no number
                query = 'query -sby name ' + show_header + ' -n *.* -u -release ' + release + ' '
                if project not in ("*","All",""):
                    # a project is selected
                    # get sub-projects
                    name, version = self.getProjectInfo(project)
                    query += '"recursive_is_member_of(cvtype=\'project\' and name=\'' + name + '\' and version=\'' + version + '\' , \'none\')" '
                query += display_attr
                executed = True
            elif project not in ("","All"):
                # No baseline, nor release selected but a project is
                query = 'query -sby name ' + show_header + ' -n *.* -u "(is_member_of(\'' + project +'\'))" '
                query += display_attr
                executed = True
            else:
                self.master_ihm.log("Please select a release or a baseline or a project.")
            if executed:
                self.master_ihm.log(" ccm " + query)
                ccm_query = 'ccm ' + query + '\n\n'
                self.master_ihm.log("List objects (directories and executable objects are discarded).")
                cmd_out = self._ccmCmd(query)
                with open(join(self.gen_dir,filename), 'w') as of:
                    if output_format == "csv":
                        header = "Release;Name;Version;Modify time;Status;Task;Task status;CR;Type\n"
                        of.write(header)
                    else:
                        of.write(ccm_query)
                    output = cmd_out.splitlines()
                    for line in output:
                        # Skip directory or relocatable objects
                        # Skip automatic tasks and components tasks
                        # Remove Baseline info at the beginning
                        if output_format == "csv":
                            if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                                of.write(line)
                                of.write("\n")
                        else:
                            if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line):
                                of.write(line)
                                of.write("\n")
        elif cmd == "HISTORY":
            filename = "log_history_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
            filename += ".csv"
            self.log.setRelease(release)
            self.log.setBaseline(baseline)
            self.log.setProject(project)
##            cid = BuildDoc("","","","","","",release,baseline,project,"SCI","","","","")
            self.log.display_attr = ' -f "%name|%version|%task|%task_synopsis|%change_request|%change_request_synopsis|%type" '
            header = ["Document","Issue","Tasks","Synopsis","CR","Synopsis"]
            self.log.tableau_items = []
            self.log.tableau_items.append(header)
            source_only = self.master_ihm.history_scope.get()
            if source_only:
                list_type_src = self.log.list_type_src_sci
                list_type_src.extend(self.log.list_type_src_hcmr)
            else:
                list_type_src = ()
            self.log.object_released = False
            self.log.object_integrate = False
            output = self.log.getArticles(list_type_src,release,baseline,project,True)
            index_src = 0
            with open(join(self.log.gen_dir,filename), 'w') as of:
                header = "File;Version;Task;Synopsis;CR;Synopsis\n"
                of.write(header)
                for line in output:
                    line = re.sub(r"<void>",r"",line)
                    m = re.match(r'(.*)\|(.*)\|(.*)\|(.*)\|(.*)\|(.*)\|(.*)',line)
                    if m:
                        result = self.log._createTblSourcesHistory(m,source_only)
                        if result:
                            index_src +=1
                            # Remove Baseline info at the beginning
                            if not re.search("(^Baseline)",line):
                                for line_csv in result:
                                    of.write(line_csv)
                                    of.write("\n")
            print "Amount of source files found: " + str(index_src)# + "\n"
            executed = True
        elif cmd == "TASKS":
            filename = "log_tasks_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
            if output_format == "csv":
                display_attr = '"%displayname;%status;%task_synopsis"'
                show_header = "-nch"
                filename += ".csv"
            else:
                display_attr = '"%displayname %status %task_synopsis"'
                show_header = "-ch"
                filename += ".txt"
            with_cr = self.master_ihm.with_cr.get()
            if baseline not in ("","All"):
                query = 'baseline -sh task ' + baseline + ' -u -f ' + display_attr + '\n'
                executed = True
            elif release not in ("","All"):
                #   -u: is not numbered
                #  -qu: query
                # -rel: release
                query = 'task -u -qu -ts all_tasks ' + show_header + ' -rel ' + release + ' -f ' + display_attr + '\n'
                executed = True
            else:
                query = 'task -u -qu -ts all_tasks ' + show_header + ' -f ' + display_attr + '\n'
                executed = True
            if executed:
                ccm_query = 'ccm ' + query + '\n'
                self.master_ihm.log(ccm_query)
                cmd_out = self._ccmCmd(query)
                with open(join(self.gen_dir,filename), 'w') as of:
                    if output_format == "csv":
                        if not with_cr:
                            header = "Task ID;Status;Synopsis\n"
                        else:
                            header = "Task ID;Task status;Task synopsis;CR ID;CR status;CR synopsis\n"
                        of.write(header)
                    else:
                        of.write(ccm_query)
                    output = cmd_out.splitlines()
                    for line in output:
                        if with_cr:
                            # Add cr information
                            mtask = re.match(r'(.*);(.*);(.*)',line)
                            # Get task ID
                            if mtask:
                                task_id = mtask.group(1)
                                task_status = mtask.group(2)
                                task_synopsis = mtask.group(3)
                                query = 'task -u -show change_request ' + task_id + ' -f "CR %problem_number;;%problem_synopsis;;%crstatus" \n'
                                ccm_query = 'ccm ' + query + '\n'
                                self.master_ihm.log(ccm_query)
                                cmd_out = self._ccmCmd(query)
                                output_cr = cmd_out.splitlines()
                                cr_id_tbl = []
                                for line_cr in output_cr:
                                    mcr = re.match(r'^CR ([0-9]*);;(.*);;(.*)$',line_cr)
                                    # Get CR ID
                                    if mcr:
                                        cr_id = mcr.group(1)
                                        cr_synopsis = mcr.group(2)
                                        cr_status = mcr.group(3)
                                        print cr_id,cr_synopsis,cr_status
                                        #  Discard CR status prefix
                                        cr_status_lite = re.sub(r'(EXCR|SyCR|ECR|SACR|HCR|SCR|BCR|PLDCR)_(.*)', r'\2', cr_status)
                                        cr_id_tbl.append([task_id,task_status,task_synopsis,cr_id,cr_status_lite,cr_synopsis])
                                    else:
                                        cr_id_tbl.append([task_id,task_status,task_synopsis,"","",""])
                                self.master_ihm.defill()
                            for task_id,task_status,task_synopsis,cr_id,cr_synopsis,cr_status in cr_id_tbl:
                                text = ""
                                text += task_id + "; " + task_status + "; " + task_synopsis + "; "
                                text += cr_id + "; " + cr_synopsis + "; " + cr_status
                                line = text
                                # Remove Baseline info at the beginning
                                if output_format == "csv":
                                    # Skip automatic tasks and components tasks
                                    if not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                                        of.write(line)
                                        of.write("\n")
                                else:
                                    # Skip automatic tasks and components tasks
                                    if not re.search("(task_automatic|component_task)",line):
                                        of.write(line)
                                        of.write("\n")
                        else:
                            # Remove Baseline info at the beginning
                            if output_format == "csv":
                                # Skip automatic tasks and components tasks
                                if not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                                    of.write(line)
                                    of.write("\n")
                            else:
                                # Skip automatic tasks and components tasks
                                if not re.search("(task_automatic|component_task)",line):
                                    of.write(line)
                                    of.write("\n")
        elif cmd == "BASELINES_DIFF":
            #
            # - objects
            # all objects that are included in the baseline are displayed. The default format is:
            #   display name, status, owner, release, create time
            #
            # - tasks
            # all objects that are included in the baseline are displayed. The default format is:
            #   id, release, assignee, create time, description
            #
            # - changes requests
            #    include details about change requests (CRs) that are partially included and fully included in the two baselines.
            #    display name, problem synopsis
            #
            query = 'baseline -compare ' + baseline_prev + ' ' + baseline_cur + ' -tasks -objects -change_requests'
            ccm_query = 'ccm ' + query + '\n'
            self.master_ihm.log(ccm_query)
            cmd_out = self._ccmCmd(query)
            filename = "log_baseline_diff_" + release_name + "_" + baseline_prev + "_vs_" + baseline_cur + "_%d.txt" % floor(time.time())
            with open(join(self.gen_dir,filename), 'w') as of:
                of.write(ccm_query)
                of.write(cmd_out)
            executed = True
        elif cmd == "BASELINES_SHOW":
            #
            # - objects
            # all objects that are included in the baseline are displayed. The default format is:
            #   display name, status, owner, release, create time
            #
            # - tasks
            # all objects that are included in the baseline are displayed. The default format is:
            #   id, release, assignee, create time, description
            #
            # - changes requests
            #    include details about change requests (CRs) that are partially included and fully included in the two baselines.
            #    display name, problem synopsis
            #
            query = 'baseline -sh objects  ' + baseline_cur
            ccm_query = 'ccm ' + query + '\n'
            self.master_ihm.log(ccm_query)
            cmd_out = self._ccmCmd(query)
            filename = "log_baseline_show_" + baseline_cur + "_%d.txt" % floor(time.time())
            with open(join(self.gen_dir,filename), 'w') as of:
                of.write(ccm_query)
                of.write(cmd_out)
            executed = True
        elif cmd == "GET_RELEASE_VS_BASELINE":
            query = "baseline -show information {:s}".format(baseline)
            ccm_query = 'ccm ' + query + '\n'
            self.master_ihm.log(ccm_query)
            cmd_out = self._ccmCmd(query)
            if cmd_out is None:
                executed = False
            else:
                filename = "log_baseline_show_" + baseline + "_%d.txt" % floor(time.time())
                with open(join(self.gen_dir,filename), 'w') as of:
                    of.write(ccm_query)
                    of.write(cmd_out)
                output = cmd_out.splitlines()
                for line in output:
                    # Attention aux espaces a supprimer
                    m = re.match(r'^  Release:( *)([^ .]*)',line)
                    if m:
                        release = m.group(2)
                        self.master_ihm.log("Associated release is: " + release)
                executed = True
        else:
            # User command
            self.master_ihm.getUserCmd()
            #cmd_txt = self.master_ihm.command_txt.get(1.0,END)
            #output = cmd_txt.splitlines()
            return
        if executed:
            self.master_ihm.log("Command executed.")
            if export:
                 # Create hyperlink
                if filename is not None:
                    #self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
                    self.master_ihm.resultGenerateCID(filename,
                                                      False,
                                                      text="COMMAND")
                else:
                    self.master_ihm.resultGenerateCID(False,
                                                      False,
                                                      text="COMMAND")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="COMMAND FAILED")
        return executed

    def _readBPROC(self,
                   bproc_filename,
                   xsl,
                   html_filename,
                   display=True):
        if bproc_filename != "":
            htmlC = HtmlConverter(bproc_filename,xsl)
            print "bproc_filename",bproc_filename
            print "html_filename",html_filename
            html_final = htmlC.toHtml(html_filename + ".html")
            if display:
                os.startfile(html_filename + ".html")

    def _checkUpdate(self):
        self.master_ihm.log("Check for new version of doCID. Please wait ...")
        new_version = self.updateCheck()
        if new_version:
            self.master_ihm.log("A new version of doCID is available: v{:s}".format(new_version))
        else:
            self.master_ihm.log("You are already running the most up-to-date version of doCID v{:s}".format(VERSION))

    def _generateCID(self,
                     release="",
                     baseline="",
                     project="",
                     implemented="",
                     item="",
                     previous_baseline="",
                     detect="",
                     cr_type="",
                     component="",
                     cr_domain = ""):
        """
        get items by invoking synergy command
        get sources by invoking synergy command
        get CR by invoking synergy command
        """
        # Create CID
        object_released = self.master_ihm.status_released
        object_integrate = self.master_ihm.status_integrate
        list_projects_set = self.master_ihm.project_set_list
        if list_projects_set == [] and not Tool.isAttributeValid(project):
            # No project set list and no project selected ?
            list_projects = self._getProjectsList_wo_ihm(release,
                                                        baseline)

        cid_type = self.master_ihm.getCIDType()

        # BuildDoc instance
        cid = BuildDoc(self.master_ihm,
                       session_started=self.getSessionStarted())
        docx_filename,exception = cid.createCID(list_projects_set,
                                                        object_released=object_released,
                                                        object_integrate=object_integrate,
                                                        cid_type=cid_type,
                                                        ccb_type=cr_domain,
                                                        item=item,
                                                        release=release,
                                                        baseline=baseline,
                                                        project=project,
                                                        target_release=implemented,
                                                        previous_baseline=previous_baseline,
                                                        detect=detect,
                                                        cr_type=cr_type,
                                                        component=component,
                                                        cr_domain=cr_domain)
        self.master_ihm.resultGenerateCID(docx_filename,
                                          exception,
                                          text="CONFIGURATION INDEX DOCUMENT")

    def _generateSQAP(self,
                    author,
                    reference,
                    revision,
                    aircraft,
                    system,
                    item):
        '''
        '''
        sqap = BuildDoc(author,reference,revision,aircraft,system,item)
        self.master_ihm.general_output_txt.insert(END, time.strftime("%H:%M:%S", time.localtime()) + " Creation doc in progress...\n")
        # Create docx
        docx_filename,exception = sqap.createSQAP()
        if docx_filename == False:
            self.master_ihm.general_output_txt.insert(END, time.strftime("%H:%M:%S", time.localtime()) + " " + exception.strerror + ", document not saved.\n")
        else:
            # Create hyperlink
            if docx_filename is not None:
                self.master_ihm.displayHyperlink("hlink",docx_filename,"Software Quality Assurance Plan in Word format.")

    def _generateCCB(self,
                     dico_parameters={"author":"",
                                      "login":"",
                                      "reference":"",
                                      "issue":"",
                                      "release":"",
                                      "baseline":"",
                                      "system":"",
                                      "item":"",
                                      "component":"",
                                      "project":"",
                                      "detect":"",
                                      "implemented":"",
                                      "cr_type":""},
                     cr_with_parent=False,
                     cr_workflow=False,
                     cr_domain=["SCR"],
                     log_on=False,
                     list_cr_for_ccb=[],
                     status_list=False,
                     ccb_time=False):
        """
        To generate a CCB report
        """
        # Get action items
        action = self.master_ihm.action
        db_exist = action.isFilenameDbExist()
        if db_exist:
            list_action_items = action.getActionItem("",1) # Only action items open
        else:
            list_action_items = []
        ccb = CCB(self.master_ihm,
                  system=dico_parameters["system"],
                  item=dico_parameters["item"],
                  detect=dico_parameters["detect"],
                  implemented=dico_parameters["implemented"],
                  cr_domain=cr_domain)
        ccb.setWorkflow(cr_workflow)
        # <NEW>
        #ccb.old_cr_workflow = ccb.get_sys_item_old_workflow(dico_parameters["system"],
        #                                                        dico_parameters["item"])
        #ccb.setDetectRelease(dico_parameters["detect"])
        #ccb.setImplRelease(dico_parameters["implemented"])
        #ccb.setDomain(cr_domain)

        ccb.setListCR(list_cr_for_ccb,
                       status_list)
        # CR list created based on list self.tableau_pr
        tableau_pr_unsorted,found_cr = ccb.getPR_CCB(cr_with_parent=cr_with_parent,
                                                      cr_type=dico_parameters["cr_type"])
        # </NEW>
        dico_former_cr_status_list = {}
        if log_on:
            if ccb_time:
                ccb_time_obj = datetime.strptime(ccb_time, '%Y/%m/%d %H:%M:%S')
            else:
                ccb_time_obj = False
            dico_cr_log = {}
            dico_cr_transition = {}
            for cr_id in list_cr_for_ccb:
                # Get transition log
                query = "query -t problem \"(problem_number='{:s}')\" -u -f \"%transition_log\"".format(cr_id)
                transi_log = self._ccmCmd(query,False)
                found_status = self.parseLog(cr_id,
                              transi_log,
                              dico_cr_transition,
                              dico_cr_log,
                              ccb_time_obj)
                if found_status:
                    #cr_id_int = int(cr_id)
                    dico_former_cr_status_list[cr_id]=found_status

        #for key,value in dico_former_cr_status_list.iteritems():
        #    print "CR ID: {:s} {:s}".format(key,value)
        docx_filename,exception = ccb.createCCB(self.list_projects,
                                                cr_domain, # deprecated to be removed
                                                list_action_items,
                                                cr_with_parent,
                                                dico_parameters,
                                                list_cr_for_ccb,
                                                status_list,
                                                ccb_time,
                                                dico_former_cr_status_list,
                                                tableau_pr_unsorted,
                                                found_cr)
        self.queue.put("RELOAD_CRLISTBOX") # action to get projects
        self.queue.put(ccb.list_change_requests)
        self.master_ihm.resultGenerateCID(docx_filename,
                                          exception,
                                          text="CHANGE CONTROL BOARD REPORT")

    def _generateDeliverySheet(self,
                              type_sds="SDS",
                              dico_tags={}
                              ):
        """
        Generate Software Delivery Sheet
        Seek .hex or .srec files to extract Part Number, Hw/Sw compatibility index and checksum
        :param type_sds:
        :param dico_tags:
        :return:
        """
        def _setOuptutFilename(template_type,dico_tags):
            """
            :return:
            """
            docx_filename = "{:s}_".format(dico_tags["system"])
            if self.item != "":
                docx_filename += "{:s}_".format(dico_tags["item"])
            if self.component != "":
                docx_filename += "{:s}_".format(dico_tags["component"])
            docx_filename += template_type + "_" + self.reference + "_%d" % floor(time.time()) + ".docx"
            self.ihm.log("Preparing " + docx_filename + " document.")
            return docx_filename

        list_projects_set = self.master_ihm.project_set_list
        if list_projects_set == [] and not Tool.isAttributeValid(dico_tags["project"]):
            # No project set list and no project selected ?
            list_projects = self._getProjectsList_wo_ihm(dico_tags["release"],
                                                         dico_tags["baseline"])
        if list_projects_set != []:
            # Projects are available in GUI
            self.master_ihm.log("Use project set list to create CID for documents",False)
            # Project set in GUI
            list_projects = self.master_ihm.project_set_list
            # List of projects from GUI
            release_text,baseline_text,project_text = self.getContext(list_projects)
        else:
            if Tool.isAttributeValid(dico_tags["project"]):
                find_sub_projects = True
                list_projects = [[dico_tags["release"],
                                  dico_tags["baseline"],
                                  dico_tags["project"]]]
                prj_name, prj_version = self.getProjectInfo(dico_tags["project"])
                self.findSubProjects(prj_name,
                                     prj_version,
                                     list_projects)
                #print "TBL",list_projects
                for sub_release,sub_baseline,sub_project in list_projects:
                    if dico_tags["project"] != sub_project:
                        self.ihm.log("Find sub project {:s}".format(sub_project))
            else:
                list_projects = [[self.release,self.baseline,""]]

        cid = BuildDoc(self.master_ihm,
                       session_started=self.getSessionStarted())
        dico_tags["eoc_id"] = ""
        tbl_bin = []
        list_found_items = []
        # Find Executable Object Code
        for release,baseline,project in list_projects:
            l_tbl_program_file = cid.getSpecificBuild(release,
                                                      baseline,
                                                      project,
                                                      filters=["BIN"],
                                                      list_found_items=list_found_items)
            self.get_eoc_infos(list_found_items,dico_tags)
            if 0==1:
                if list_found_items != []:
                    print "list_found_items",list_found_items
                    for object in list_found_items:
                        m = re.match(r'^(.*)\.(.*)-(.*):(.*):([0-9]*)$',object)
                        if m:
                            ext = m.group(2)
                            print "EXT",ext
                            if (ext == "hex") or (ext == "srec"):
                                eoc_filename = "eoc" + "_%d" % floor(time.time()) + "." + ext
                                self.catEOC(object,eoc_filename)
                                dico_addr = self.getEOCAddress()
                                hw_sw_compatibility,pn,checksum = self._readEOC(join("result",eoc_filename),dico_addr)
                                dico_tags["part_number"] = pn # Ex: ECE3E-A338-0501
                                dico_tags["eoc_id"] =  re.sub(r'ECE[A-Z0-9]{2}-A([0-9]{3})-([0-9]{4})',r'A\1L\2',pn)
                                dico_tags["checksum"] = checksum # Ex: 0x6b62
                                dico_tags["hw_sw_compatibility"] = hw_sw_compatibility # Ex 0x100
                                break
            tbl_bin.extend(l_tbl_program_file)
        print "tbl_bin",tbl_bin
        pn = self.getComponentPartNumber(dico_tags["component"])
        list_tags = {
                    'CI_ID':{'type':'str','text':pn,'fmt':{}},
                    'REFERENCE':{'type':'str','text':dico_tags["reference"],'fmt':{}},
                    'ISSUE':{'type':'str','text':dico_tags["issue"],'fmt':{}},
                    'ITEM':{'type':'str','text':dico_tags["item"],'fmt':{}},
                    'COMPONENT':{'type':'str','text':dico_tags["component"],'fmt':{}},
                    'DATE':{'type':'str','text':time.strftime("%d %b %Y", time.localtime()),'fmt':{}},
                    'PROJECT':{'type':'str','text':dico_tags["project"],'fmt':{}},
                    'RELEASE':{'type':'str','text':dico_tags["release"],'fmt':{}},
                    'BASELINE':{'type':'str','text':dico_tags["baseline"],'fmt':{}},
                    'WRITER':{'type':'str','text':dico_tags["author"],'fmt':{}},
                    'PART_NUMBER':{'type':'str','text':dico_tags["part_number"],'fmt':{}},
                    'CHECKSUM':{'type':'str','text':dico_tags["checksum"],'fmt':{}},
                    'PROTOCOL_COMPAT':{'type':'str','text':"",'fmt':{}},
                    'DATA_COMPAT':{'type':'str','text':"",'fmt':{}},
                    'HW_COMPAT':{'type':'str','text':dico_tags["hw_sw_compatibility"],'fmt':{}},
                    'EOC_ID':{'type':'str','text':dico_tags["eoc_id"],'fmt':{}}}
        template_dir = join(os.path.dirname("."), 'template')
        template_name = self.getOptions("Template","SDS")
        template = join(template_dir, template_name)
        # Prepare output file
        docx_filename = _setOuptutFilename("SDS",dico_tags)
        self.ihm.docx_filename = docx_filename
        docx_filename,exception = self._createDico2Word(list_tags,
                                                             template,
                                                             docx_filename)
        if not docx_filename:
            self.master_ihm.log(exception + ": document not saved.")
        else:
            try:
                self.master_ihm.cid_word_img_can.itemconfigure(self.master_ihm.cid_word_img,state='normal')
                self.master_ihm.success.config(fg='magenta',bg = 'green',text="SOFTWARE DELIVERY SHEET GENERATION SUCCEEDED")
            except AttributeError:
                pass
            self.master_ihm.displayHyperlink("hlink",docx_filename,"Software Delivery Sheet in Word format.")
        # Set scrollbar at the bottom
        self.master_ihm.defill()

    def _generateReviewReport(self,
                              review_number,
                              empty=False):
        """
        :param review_number:
        :param empty:
        :return:
        """
        # Create docx
        try:
            self.master_ihm.success.config(fg='red',bg = 'yellow',text="REVIEW REPORT GENERATION IN PROGRESS")
        except AttributeError:
            pass
        dico = self.master_ihm.getParameters()
        cr_type = dico["cr_type"]
        release = dico["release"]
        baseline = dico["baseline"]
        project = dico["project"]
        if project in ("","None","All",None):
            project = ""
        if baseline in ("","None","All",None):
            baseline = ""
        review_qams_id = self.master_ihm.review_qams_id
        conformity_level = self.master_ihm.var_conformity.get()
        project_list = []
        if not empty:
            if self.master_ihm.project_set_list == []:
                project_list.append([release,
                                     baseline,
                                     project])
            else:
                project_list = self.master_ihm.project_set_list
        else:
            pass
        print "DICO before Review init",dico
        review = Review(review_number,
                        detect_release=dico["detect"],
                        impl_release=dico["implemented"],
                        session_started=self.session_started,
                        project_list=project_list,
                        author=dico["author"],
                        system=dico["system"],
                        item=dico["item"],
                        component=dico["component"],
                        part_number=dico["part_number"],
                        checksum=dico["checksum"],
                        reference=dico["reference"],
                        issue=dico["issue"],
                        review_qams_id=review_qams_id,
                        conformity_level=conformity_level,
                        cr_type = cr_type,
                        sw_level = dico["dal"],
                        ihm=self.master_ihm)
        review_name = review.getName(review_number)
        self.master_ihm.log(("Creation {:s} review report in progress...").format(review_name))
        docx_filename,exception = review.createReviewReport(empty,
                                                            review_number,
                                                            detect_release=dico["detect"],
                                                            impl_release=dico["implemented"])
        #old_review.docx_filename = docx_filename
        if not docx_filename:
            self.master_ihm.log(exception + ": document not saved.")
        else:
            try:
                self.master_ihm.success.config(fg='magenta',bg = 'green',text="REVIEW REPORT GENERATION SUCCEEDED")
            except AttributeError:
                pass
            self.master_ihm.displayHyperlink("hlink",docx_filename,"Review report in Word format.")
        # Set scrollbar at the bottom
        self.master_ihm.defill()

    def _generateDocument(self,template_key,list_tags):
        """
        generat generic document with tag list input
        """
        self.master_ihm.log("Creation document in progress...")
        # Create docx
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="DOCUMENT GENERATION IN PROGRESS")
        docx_filename,exception = generic_doc.create(list_tags,
                                                    template_key)
        if not docx_filename:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="DOCUMENT GENERATION FAILED")
            self.master_ihm.log(exception.strerror + ", document not saved.")
        else:
            self.master_ihm.success.config(fg='magenta',bg = 'green',text="DOCUMENT GENERATION SUCCEEDED")
            self.master_ihm.displayHyperlink("hlink",docx_filename,"Document in Word format.")
        # Set scrollbar at the bottom
        self.master_ihm.defill()

    def _closeSession(self):
        '''
        Close session
        '''
        global session_started
        query = "stop"
        self.master_ihm.log("ccm " + query)
        stdout = self._ccmCmd(query)
        # Set scrollbar at the bottom
        self.master_ihm.general_output_txt.see(END)

    def _getSessionStatus(self):
        '''
        Retrieve database used
        '''
        global session_started
        query = "status"
        self.master_ihm.log("ccm " + query)
        stdout = self._ccmCmd(query)
        # Set scrollbar at the bottom
        try:
            self.master_ihm.general_output_txt.see(END)
        except AttributeError:
            pass
        output = stdout.splitlines()
        for line in output:
            m = re.search(r'Database:(.*)',line)
            if m:
                database = m.group(1)
                self.master_ihm.log("Database used is:" + database)
                return_code = database
                break;
            else:
                return_code = False
        if not return_code:
            print "No Synergy database found"
        return return_code

    def _checkISCmd(self,
                    dirname_upper="",
                    dirname_req="",
                    filename_is="",
                    component="",
                    hlr_selected=False):
        """
        This function checks Inspection Sheet document for specification
        :param dirname_upper:
        :param dirname_req:
        :param filename_is:
        :param hlr_selected:
        :return:
        """

        skip_change_synergy_var = self.master_ihm.skip_change_synergy_var.get()
        #print "SKIP:",skip_change_synergy_var
        check_is = CheckIS(dirname_req,
                           hlr_selected = hlr_selected,
                           general_output_txt = self.master_ihm.general_output_txt,
                           session_started=True)
        if hlr_selected:
            tbl_type=["SWRD","PLDRD"]
            check_is.openLog("RD")
        else:
            tbl_type=("SWDD",)
            check_is.openLog("SWDD")
        #print "_checkISCmd:filename_is",filename_is
        doc_upper,doc_inspected,filename_is_short = check_is.checkISForSpec(filename_is = filename_is,
                                                                  dirname_req = dirname_req,
                                                                  dirname_upper = dirname_upper,
                                                                  type = tbl_type,
                                                                  skip_change_synergy_var=skip_change_synergy_var,
                                                                  component=component)

        # Export results of analysis in an Excel workbook
        report_filename = check_is.export(doc_upper = doc_upper,
                                          doc_inspected = doc_inspected,
                                          filename_is = filename_is_short)
        check_is.closeLog()
        if check_is.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink1",check_is.log_filename)

        if report_filename is not None:
            self.master_ihm.resultGenerateCID(filename,
                                                False,
                                                text="INSPECTION CHECK")
        if 0==1:
            if report_filename is not None:
                self.master_ihm.displayHyperlink("hlink2",report_filename,"IS check report")
            if doc_inspected:
                self.master_ihm.success.config(fg='magenta',bg = 'green',text="INSPECTION CHECK SUCCEEDED")
            else:
                self.master_ihm.success.config(fg='yellow',bg = 'red',text="INSPECTION CHECK FAILED")

    def _checkISDocCmd(self,filename_is,verif_issue_cr_process_start):
        """
        This function checks Inspection Sheet document for any other documents
        :param filename_is:
        :param verif_issue_cr_process_start:
        :return:
        """
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="INSPECTION CHECK IN PROGRESS ...")
        skip_change_synergy_var = self.master_ihm.skip_change_synergy_var.get()
        check_is = CheckIS("",
                           general_output_txt = self.master_ihm.general_output_txt)
        check_is.openLog("Generic")
        result = check_is.CheckISGeneric(filename_is,skip_change_synergy_var,verif_issue_cr_process_start)
        check_is.logErrors()
        check_is.logWarnings()
        check_is.closeLog()
        if check_is.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink",check_is.log_filename)
        if result:
            self.master_ihm.success.config(fg='magenta',bg = 'green',text="INSPECTION CHECK SUCCEEDED")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="INSPECTION CHECK FAILED")

    def _getChapterDialogHLR(self,filename):
        CheckLLR.getChapterReq(filename)

    def _exportIS(self,
                  dirname_req="",
                  dirname_upper="",
                  hlr_selected=False,
                  reference="",
                  issue="",
                  release="",
                  hsid_dirname=""):

        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        export_is = CheckIS(dirname_req,
                           hlr_selected = hlr_selected,
                           general_output_txt = self.master_ihm.general_output_txt,
                           session_started=True)

        if not hlr_selected:
            export_is.openLog("SWDD")
            export_is.getHSID(hsid_dirname)
            spec_ref = "SWDD_{:s}-{:s}".format(reference,issue)
            # Extract requirements from SWDD
            export_is.extract(dirname_req,("SWDD",))
        else:
            export_is.openLog("SWRD")
            spec_ref = "SWRD_{:s}-{:s}".format(reference,issue)
            # Extract requirements from SWRD
            export_is.extract(dirname_req,("SWRD",))
        export_is.closeLog()
        print "tbl_file_llr_wo_del",export_is.tbl_file_llr_wo_del
        print "tbl_list_llr",export_is.tbl_list_llr
        # Extract requirements from upper specifications.
        upper = CheckLLR(dirname_upper,hlr_selected=True)
        if hlr_selected:
            upper.openLog("SSCS")
            list_upper = upper.getListUpper()
            upper.extract(dirname_upper,
                          type=list_upper)
        else:
            upper.openLog("SWRD")
            list_upper = ("SWRD",)
            upper.extract(dirname_upper,
                          type=list_upper)
        #print "list_upper",list_upper
        #print "upper.tbl_list_llr",upper.tbl_list_llr
        #for x in upper.tbl_list_llr:
        #    print "upper:",x
        # Create excel workbook

        # CR list
        dico_parameters = self.master_ihm.getParameters()
        ccb = CCB(self.master_ihm)
        dico_tableau_pr = {"all":[],
                           "open":[],
                           "closed":[]}
        ccb.getPR(dico_tableau_pr,
                   dico_parameters["detect"],
                   dico_parameters["implemented"],
                   dico_parameters["cr_type"],
                   False)

        for pr,list in dico_tableau_pr.iteritems():
            print pr,list
        print "upper.tbl_list_llr",upper.tbl_list_llr
        filename_is = export_is.exportIS(spec_ref,
                                         reference=reference,
                                         issue=issue,
                                         release=release,
                                         reviewer_name = self.master_ihm.reviewer_name_entry.get(),
                                         default_status = self.master_ihm.default_status,
                                         author=dico_parameters["author"],
                                         project=dico_parameters["system"],
                                         dico_upper=upper.tbl_list_llr,
                                         tbl_cr=dico_tableau_pr["all"])
        upper.closeLog()

        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        if filename_is is not None:
            self.master_ihm.displayHyperlink("hlink",filename_is,"Inspection Sheet created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="INSPECTION SHEET EXPORT SUCCEEDED")

    def _checkLLRCmd(self,
                     dirname="",
                     hlr_selected=False,
                     list_spec=("SWRD","PLDRD"),
                     hsid_dirname=""):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="SPECIFICATION CHECK IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        llr = CheckLLR(dirname,
                       hlr_selected = hlr_selected,
                       general_output_txt = self.master_ihm.general_output_txt)
        if hlr_selected:
            llr.openLog("SwRD")
        else:
            llr.openLog("SwDD")
            llr.getHSID(hsid_dirname)
        attr_check_filename,file_check_filename = llr.extract(dirname,list_spec)
        llr.logErrors()
        llr.logWarnings()
        llr.closeLog()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        if attr_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink1",attr_check_filename,"List of requirements with attributes.")
        if file_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink3",attr_check_filename,"List of files with amount of requirenents per file.")
        if llr.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink2",llr.log_filename,"Log created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="SPECIFICATION CHECK SUCCEEDED")

    def _checkUpperCmd(self,dirname):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="SPECIFICATION CHECK IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        upper = CheckLLR(dirname,hlr_selected=True)
        upper.openLog("SSCS")
        list_upper = upper.getListUpper()
        attr_check_filename,file_check_filename = upper.extract(dirname,list_upper)
        upper.getAllocation()
        #print "TEST",upper.dico_alloc_vs_req
        upper.logErrors()
        upper.logWarnings()
        upper.closeLog()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        self.master_ihm.log("{:d} requirements found.".format(upper.nb_reqs))
        if attr_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink1",attr_check_filename,"List of requirements with attributes.")
        if file_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink3",attr_check_filename,"List of files with amount of requirenents per file.")
        if upper.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink2",upper.log_filename)
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="SPECIFICATION CHECK SUCCEEDED")

    def _genLLRDerivedCmd(self,dirname,tbl_type=("SWDD",)):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="DERIVED REQUIREMENTS EXPORT IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        hlr = Derived(dirname,
                      hlr_selected=False,
                      general_output_txt = self.master_ihm.general_output_txt)
        hlr.listDir(dirname,tbl_type)
        hlr.invert()
        hlr.countDerived()
        filename = hlr.export()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))

        if filename is not None:
            self.master_ihm.displayHyperlink("hlink",filename,"xlsx document created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="DERIVED REQUIREMENTS EXPORT SUCCEEDED")

    def _genHLRDerivedCmd(self,dirname,tbl_type=["SWRD","PLDRD"]):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="DERIVED REQUIREMENTS EXPORT IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        hlr = Derived(dirname,
                      hlr_selected=True,
                      general_output_txt = self.master_ihm.general_output_txt)
        hlr.listDir(dirname,tbl_type)
        hlr.invert()
        if "SWRD" in tbl_type:
            hlr.countDerived()
        else:
            hlr.countDerived("SSCS")
        filename = hlr.export()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))

        if filename is not None:
            self.master_ihm.displayHyperlink("hlink",filename,"xlsx document created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="DERIVED REQUIREMENTS EXPORT SUCCEEDED")

    @staticmethod
    def extractCR(tbl_cr):
        cr_id = tbl_cr[0]
        cr_synopsis = tbl_cr[2]
        cr_status = tbl_cr[3]
        return cr_id,cr_synopsis,cr_status

    def cleanImpactAnalysis(self,impact_analysis):
        #impact_analysis = re.sub(r", ?$",r"",impact_analysis)
        import html2text
        impact_analysis_plain_txt = html2text.html2text(Tool.removeNonAscii(impact_analysis))
        #impact_analysis_plain_txt = re.sub(r"\r",r" ",impact_analysis_plain_txt)
        #impact_analysis_plain_txt = re.sub(r"\n",r" ",impact_analysis_plain_txt)
        #print "impact_analysis_plain_txt",impact_analysis_plain_txt
        return impact_analysis_plain_txt

    def export(self,
               tbl_cr,
               for_review_on=False):
        """
        export CRs list in Excel file
        :param tbl_cr:
        :param cr_type:
        :param for_review_on:
        :return:
        """
        sheet_name = 'Change Requests'
        wb = load_workbook(filename = join('template',self.export_cr_list_filename))
        if wb is not None:
            ws = wb.get_sheet_by_name(name = sheet_name)
            filename = None
            if ws is not None:
                CheckLLR.putLogo(ws)
                style_border = Style(border=Border(
                    left=Side(border_style=BORDER_THIN),
                    right=Side(border_style=BORDER_THIN),
                    top=Side(border_style=BORDER_THIN),
                    bottom=Side(border_style=BORDER_THIN)),
                                     alignment=Alignment(wrap_text=True,shrink_to_fit=True))
                row = 9
                if not for_review_on:
                    for cr in tbl_cr:
                        # ID
                        cr_id = cr[0]
                        # Patch to get CR domain from CR status
                        cr_domain = self.getStatusPrefix(cr[3])
                        cr[0] = "{:s} {:s}".format(cr_domain,cr[0])
                        # Synopsis
                        cr[2] = self.replaceNonASCII(cr[2])
                        # Status
                        cr[3] = self.removeStatusPrefix(cr[3])

                        # Impact analysis
                        cr[9] = self.cleanImpactAnalysis(cr[9])
                        #print cr
                        #hyperlink = "http://spar-syner1.in.com:8600/change/PTweb?ACTION_FLAG=frameset_form&TEMPLATE_FLAG=ProblemReportView&database=%2Fusr%2Flocal%2Fccmdb%2Fdb_sms_pds&role=User&problem_number={:s}".format(cr_id)
                        #hyperlink = "http://www.lemonde.fr"
                        #CheckLLR.setCell(ws,cr,row,1,style_border)
                        #CheckLLR.setHyperlink(ws,row,1,hyperlink)
                        for col_idx in range(1,15):
                            CheckLLR.setCell(ws,cr,row,col_idx,style_border)
                        row += 1
                    # Autofilter
                    ws.auto_filter.ref = "A8:N8"
                    CheckLLR.set_border(ws, "K7:N7")
                else:
                    for cr in tbl_cr:
                        for col_idx in range(1,6):
                            CheckLLR.setCell(ws,cr,row,col_idx,style_border)
                        row += 1
                    # Autofilter
                    ws.auto_filter.ref = "A8:N8"
                    CheckLLR.set_border(ws, "K7:N7")
                # save the file
                filename = "Change_Requests_List_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))
            else:
                print "WorkSheet \"{:s}\" access failed.".format(sheet_name)
                ws_found = wb.get_sheet_names()
                print "Found:",ws_found
        else:
            print "WorkBook \"{:s}\" access failed.".format(self.export_cr_list_filename)
        return filename

    @staticmethod
    def getTransition(line):
        match_transition = re.match(r'^_TRANSITION_ (Submitted to|Transitioned to) (.*) by (.*) on ([0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2})',line)
        if match_transition:
            transition = ThreadQuery.removeStatusPrefix(match_transition.group(2))
            date_change = match_transition.group(4)
            datetime_obj = datetime.strptime(date_change, '%Y/%m/%d %H:%M:%S')
            date = datetime_obj.strftime('%A %d %b %Y')
            time = datetime_obj.strftime('%H:%M:%S')
        else:
            transition = False
            date = False
            time = False
            datetime_obj = False
        return transition,date,time,datetime_obj

    def parseLog(self,
                 cr_id,
                 transi_log,
                 dico_cr_transition,
                 dico_cr_log,
                 ccb_time_obj=False):
        """

        :param cr_id:
        :param transi_log:
        :param dico_cr_transition:
        :param dico_cr_log:
        :param ccb_time:
        :return: found_status
        """
        # Replace FS and RS characters
        char = {r'\x1c':'_TRANSITION_ ',
                r'\x1e':'_UPDATE_ '}
        for before, after in char.iteritems():
            transi_log = re.sub(before,after,transi_log)
        if transi_log is not None:
            transi_log_filtered = self.replaceNonASCII(transi_log)
            #transi_log_filter.decode('latin1') #filter(string.printable[:-5].__contains__,transi_log_filter)
        else:
            transi_log_filtered = transi_log
        #transi_log_filtered = self._filterASCII(transi_log)
        dico_cr_log[cr_id] = transi_log_filtered
        tbl_log = transi_log_filtered.splitlines()
        list_transitions = []
        check_comment = False
        transition = False
        found_status = False
        for line in tbl_log:
            if transition:
                if check_comment:
                    # Get comment for transition
                    list_transitions.append((transition,date,time,line))
                    check_comment = False
                else:
                    list_transitions.append((transition,date,time,""))
                transition = False
            else:
                transition,date,time,datetime_obj = self.getTransition(line)
                if datetime_obj and ccb_time_obj:
                    if datetime_obj > ccb_time_obj:
                        if not found_status:
                            found_status = transition
                    elif datetime_obj < ccb_time_obj:
                        found_status = transition
                    else:
                        pass
                if transition in ("Under_Modification",
                                  "Closed",
                                  "Rejected",
                                  "Postponed",
                                  "Complementary_Analysis"):
                    # According to chapter 3.3.3.3 of SCMP, CCB minutes reference shall be documented on Comments field
                    # Missing "incomplete analysis" transition
                    check_comment = True
        dico_cr_transition[cr_id] = list_transitions
        return found_status

    def _preview_CR_Query(self):
        self.master_ihm.preview_CR_Query()

    def _getCR(self,
               baseline="",
               extension=True,
               for_review_on=False,
               cr_with_parent = False,
               log_on = False,
               component_type="",
               detected_on="",
               implemented_for="",
               old_cr_workflow=False,
               ccb_time=False):
        '''
            List CR
            Generate an Excel file at the end
            get
                variables
                    previous_release,
                    impl_release,
                    baseline,
                    project
                    attribute ? Encore utilise ? pas sur.
                methods
                    getTypeWorkflow
                    cr_for_review_var
            from ThreadQuery <=== Interface class
            set
                methods
                    setPreviousRelease
                    setRelease
                    setBaseline
                    setProject
            to BuildDoc

            Note: baseline is used only fo A/C standard
        '''
        # Create CR list
        if ccb_time:
            ccb_time_obj = datetime.strptime(ccb_time, '%Y/%m/%d %H:%M:%S')
        else:
            ccb_time_obj = False
        output = ""
        log_filename = "log_list_crs_%d" % floor(time.time()) + ".txt"
        # Domain
        #self.ccb_type = cr_type

        if Tool.isAttributeValid(baseline):
            # get standard
            list_sub_std = []
            if self.master_ihm.dico_std.has_key(baseline):
                #
                # Cette partie ne marche plus, a checker
                #
                condition = '"(cvtype=\'problem\') '
                filter_cr = ""
                list_sub_std = self.master_ihm.dico_std[baseline]
                find_std = False
                implemented = ""
                num = 0
                if self.master_ihm.dico_list_std.has_key(baseline):
                    # Est-ce un standard avion ou un sous-standard projet ?
                    pass
                else:
                    delta_implemented,find_std = self.createCrImplemented(baseline,find_std,filter_cr)
                    implemented += delta_implemented
                for sub_std in list_sub_std:
                        delta_implemented,find_std = self.createCrImplemented(sub_std,find_std,filter_cr)
                        implemented += delta_implemented
                if find_std == True:
                    implemented +=  ') '
                condition += implemented
            else:
                pass

        condition,detect_attribut = self.master_ihm._createConditionStatus(detect_release = detected_on,
                                                                           impl_release = implemented_for,
                                                                           cr_type = component_type)

        classification = CCB.getClassif(old_cr_workflow)
        if for_review_on:
            attributes = '-f "<cell>%problem_number</cell>' \
                         '<cell>%problem_synopsis</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>%crstatus</cell>' \
                         '<cell>%void</cell>"'.format(classification)
        else:
            #if not old_cr_workflow:
            # New problem report workflow
            implementation_baseline_f = "%CR_implementation_baseline"
            # new with tags
            detect_attribut_tag = re.sub(r";","</cell><cell>",detect_attribut)
            attributes = '-f "<cell>%problem_number</cell>' \
                         '<cell>%CR_request_type</cell>' \
                         '<cell>%problem_synopsis</cell>' \
                         '<cell>%crstatus</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>%modify_time</cell>' \
                         '<cell>%impact_analysis</cell>"'.format(classification,detect_attribut_tag,implementation_baseline_f)
        # query with no numbering of the line and sorted by problem_number
        query = "query -u -sby problem_number {:s} {:s} ".format(condition,attributes)
        self.master_ihm.log('ccm ' + query)
        # remove \n
        text = re.sub(r"\n",r"",query)
        stdout,stderr = self.ccm_query(text,'Get CRs for CCB minutes')
        list_change_requests = []
        tbl_cr_export = []
        dico_cr_log = {}
        dico_cr_transition = {}
        filename = None
        if stdout != "":
            stdout = self.replaceBeacon(stdout)
            char = {r"<br ?\/>":r", ",
                    r"\r\n":r"\n",
                    r"\x1E":"----------\n",
                    r"\x1C":r"\n",
                    r"<void>":r""}
            for before,after in char.iteritems():
                stdout = re.sub(before,after,stdout)
            result = stdout
            #self.master_ihm.log(result)
            output = result.splitlines()
            for line in output:
                cr_decod = self._parseCRCell(line)
                cr_id,cr_synopsis,cr_status = self.extractCR(cr_decod)
                #  Used to fill UI CR list box
                CCB.createCRlist(cr_id,
                                 cr_synopsis,
                                 list_change_requests)
                # For CLI
                #print line
                if cr_with_parent:
                    tbl_parent_cr_id = self._getParentCR(cr_id)
                    if tbl_parent_cr_id:
                        # Get parent ID pieces of information
                        found_parent_cr_info = []
                        found_parent_cr_status = []
                        found_parent_cr_synopsis = []
                        found_parent_cr_implemented_for = []
                        nb_parent_cr = len(tbl_parent_cr_id)
                        index_parent_cr = 1
                        for parent_cr_id in tbl_parent_cr_id:
                            parent_cr = self._getParentInfo(parent_cr_id)
                            cr_info = ""
                            parent_cr_status = ""
                            cr_synopsis = ""
                            if parent_cr:
                                parent_decod = self._parseCRParent(parent_cr)
                                # Parent CR;Parent CR status;Parent CR synopsis
                                cr_info = parent_decod[0] + " " + parent_decod[1] + " " + parent_decod[2]
                                parent_cr_status = CCB.discardCRPrefix(parent_decod[3])
                                cr_synopsis = Tool.replaceNonASCII(parent_decod[4])
                                cr_implemented_for = parent_decod[5]
                                found_parent_cr_info.append(cr_info)
                                found_parent_cr_status.append(parent_cr_status)
                                found_parent_cr_synopsis.append(cr_synopsis)
                                found_parent_cr_implemented_for.append(cr_implemented_for)
                        found_parent_cr_info_str = ", ".join(map(str, found_parent_cr_info))
                        found_parent_cr_status_str = ", ".join(map(str, found_parent_cr_status))
                        found_parent_cr_synopsis_str = ", ".join(map(str, found_parent_cr_synopsis))
                        found_parent_cr_implemented_for_str = ", ".join(map(str, found_parent_cr_implemented_for))
                        cr_decod.extend([found_parent_cr_info_str,
                                         found_parent_cr_status_str,
                                         found_parent_cr_synopsis_str,
                                         found_parent_cr_implemented_for_str])
                    else:
                        cr_decod.extend(["","","",""])
                else:
                    cr_decod.extend(["","","",""])
                tbl_cr_export.append(cr_decod)
                if log_on:
                    # Get transition log
                    query = "query -t problem \"(problem_number='{:s}')\" -u -f \"%transition_log\"".format(cr_id)
                    transi_log = self._ccmCmd(query,False)


                    found_status = self.parseLog(cr_id,
                                  transi_log,
                                  dico_cr_transition,
                                  dico_cr_log,
                                  ccb_time_obj)
                    if found_status:
                        print "CR ID: {:s} {:s} <-- {:s}".format(cr_id,found_status,cr_status)

            # end loop for CR parsing
            # Create Excel file with CRs listing
            filename = self.export(tbl_cr_export,
                                   for_review_on=for_review_on)
        list_change_requests.sort()
        self.master_ihm.reloadCR_ListBox(list_change_requests)
        #self.queue.put("RELOAD_CRLISTBOX") # action to get projects
        #self.queue.put(list_change_requests)
        if stderr:
            print time.strftime("%H:%M:%S", time.localtime()) + " " + stderr
             # remove \r
            result = re.sub(r"\r\n",r"\n",stderr)
            self.master_ihm.log(result)
        with open(join(self.gen_dir,log_filename), 'w') as of:
            ccm_query = 'ccm ' + query + '\n\n'
            of.write(ccm_query)
            for cr_id,log in dico_cr_log.iteritems():
                txt = "Full log for {:4s} {:5s}:\n".format(cr_type,cr_id)
                of.write(txt)
                of.write("-----------------------\n\n")
                txt = "{:s}\n".format(log)
                of.write(txt)
            of.write("\n\n--------------------------------------------------------------------------------------------------------------\n\n")
            for cr_id,transitions in dico_cr_transition.iteritems():
                txt = "Transitions timeline for {:4s} {:5s}:\n".format(cr_type,cr_id)
                of.write(txt)
                of.write("-----------------------------------\n\n")
                for transition,date,hour,comment in transitions:
                    if comment != "":
                        txt = "   Status set to {:20s} on {:25s} at {:15s} with comment: {:s}\n".format(transition,date,hour,comment)
                    else:
                        txt = "   Status set to {:20s} on {:25s} at {:15s} with no comment.\n".format(transition,date,hour)
                    of.write(txt)
                of.write("\n")
            #of.write(result)
        #self.master_ihm.log("Command executed.")
        self.master_ihm.resultGenerateCID(filename,None,"CHANGE REQUEST LISTING")
        try:
            if 0==1:
                if filename is not None:
                    self.master_ihm.docx_filename = filename
                    self.master_ihm.ccb_word_img_can.itemconfigure(self.master_ihm.ccb_word_img,state='normal')
                    self.master_ihm.success.config(fg='magenta',bg = 'green',text="CHANGE REQUEST LISTING SUCCEEDED")
                    self.master_ihm.displayHyperlink("hlink1",filename,"List of CR in Excel file.")
                else:
                    self.master_ihm.success.config(fg='yellow',bg = 'red',text="CHANGE REQUEST LISTING  FAILED")
            if log_filename is not None:
                self.master_ihm.displayHyperlink("hlink2",log_filename,"Log of CR and transitions summary.")
        except AttributeError:
            pass
        # Set scrollbar at the bottom
        #self.master_ihm.defill()
        # For debug purpose
        return output

    def run(self):
        # sleep to enables the GUI to finish its setting
        import time
##        print time.strftime("%H:%M:%S", time.localtime()) + " Start thread " + self.name_id + "\n"
        time.sleep(2)
        self.periodicCall()

    def stop(self):
##        print time.strftime("%H:%M:%S", time.localtime()) + " Stop thread " + self.name_id + "\n"
        self.terminated = True

if __name__ == '__main__':
    is_excel = ThreadQuery()
    is_excel.checkISCmd()
    exit()
    cr = "818;SCR_Closed;SCR"
    m = re.match(r'^([0-9]*);(EXCR|SyCR|ECR|SACR|HCR|SCR|BCR|PLDCR)_(.*);(.*)',cr)
    if m:
        cr_id = m.group(1)
        cr_status = m.group(3)
        cr_domain = m.group(4)
        print "TEST:",cr_id
    test_str = "SMS_EPDS_SPI_ICD_ET3532_S.pdf issue 3 for release SW_ENM/05"
    print len(test_str)
    print test_str
    result = test_str.ljust(80,"*")
    print result